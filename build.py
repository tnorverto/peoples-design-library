"""Build The People's Design Library website from the Google Sheet.

Downloads the spreadsheet as .xlsx, extracts every hyperlink with its
section/subheader context, cleans it, and injects it into shell.html
to produce index.html.

Usage:  python build.py            (downloads from SHEET_ID)
        python build.py file.xlsx  (uses a local file instead)
"""
import json, re, sys, urllib.request
from urllib.parse import urlparse
from openpyxl import load_workbook

SHEET_ID = "13GStMRQfbn5glWVkUPqFtW1oovyKhMDdRKD3m5cstBg"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# (sheet tab name, display name, call-number code, spine color)
COLLECTIONS = [
    ("GRAPHIC  GENERAL DESIGN", "Graphic & General Design", "GD", "#2E6E7E"),
    ("VIDEO  ANIMATION  SOUND", "Video, Animation & Sound", "VA", "#7A4E8C"),
    ("ARCHITECTURE  3D", "Architecture & 3D", "AR", "#B07A2A"),
    ("AI  SOFTWARE  PLUGINS", "AI, Software & Plugins", "AI", "#3D6B35"),
    ("TUTORIALS", "Tutorials", "TU", "#34518F"),
    ("EXTRAS", "Extras", "EX", "#A8502F"),
]
SKIP_SHEETS = {"MENU", "COMMUNITY"}

SECTION_RENAME = {
    "SOUND DESIGN / ENGEENEERING": "SOUND DESIGN / ENGINEERING",
    "SOME EXTRAS...": "USEFUL SITES & MISC",
}
# Sections whose header cell is a long note get renamed by prefix match
SECTION_PREFIX_RENAME = [
    ("Search 3d models by pic", "3D MODELS"),
    ("1\U0001FA99", "A.I. TOOLS"),  # the AI token legend row
]

FIRE = re.compile("\U0001F525+")
WS = re.compile(r"\s+")


def download(path="sheet.xlsx"):
    print("Downloading sheet…")
    req = urllib.request.Request(EXPORT_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=120) as r, open(path, "wb") as f:
        f.write(r.read())
    return path


def is_header_row(ws, row):
    for mc in ws.merged_cells.ranges:
        if mc.min_row == row and mc.max_row == row and mc.min_col <= 2 and mc.max_col > mc.min_col:
            return True
    return False


def clean_section(s):
    if s is None:
        return "General"
    s = SECTION_RENAME.get(s, s)
    for prefix, name in SECTION_PREFIX_RENAME:
        if s.startswith(prefix):
            s = name
    s = s.replace("\U0001F3F4\u200d\u2620\ufe0f", "").replace("\u200d", "")
    return WS.sub(" ", s).strip().rstrip(":").strip()


def clean_name(n):
    pick = bool(FIRE.search(n))
    n = FIRE.sub("", n).replace("\U0001F3F4\u200d\u2620\ufe0f", "").replace("\u200d", "")
    n = WS.sub(" ", n).strip(" -–—:").strip()
    if len(n) > 110:
        n = n[:110].rsplit(" ", 1)[0] + "…"
    return n, pick


def domain(u):
    try:
        d = urlparse(u.strip()).netloc.lower()
        return d[4:] if d.startswith("www.") else d
    except Exception:
        return ""


def parse_default(sheet_name, ws):
    """Parallel column lists; bold non-linked cells are subheaders."""
    entries, current_section, subheaders = [], None, {}
    for r in range(1, ws.max_row + 1):
        if is_header_row(ws, r):
            val = ws.cell(row=r, column=2).value
            if val:
                current_section, subheaders = str(val).strip(), {}
            continue
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if not val:
                continue
            link = cell.hyperlink.target if cell.hyperlink else None
            if link:
                entries.append((sheet_name, current_section, subheaders.get(c), val, link))
            elif cell.font and cell.font.bold:
                subheaders[c] = val
    return entries


def parse_tutorials(sheet_name, ws):
    """Rows of: B=topic, C=title, D=link."""
    entries, current_section = [], None
    for r in range(1, ws.max_row + 1):
        if is_header_row(ws, r):
            val = ws.cell(row=r, column=2).value
            if val:
                current_section = str(val).strip()
            continue
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        dcell = ws.cell(row=r, column=4)
        link = dcell.hyperlink.target if dcell.hyperlink else None
        if link and c:
            entries.append((sheet_name, current_section, str(b).strip() if b else None, str(c).strip(), link))
    return entries


def build(xlsx_path):
    wb = load_workbook(xlsx_path)
    col_idx = {k: i for i, (k, *_r) in enumerate(COLLECTIONS)}
    raw = []
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        if name not in col_idx:
            print(f"  ! Unknown tab skipped: {name!r} (add it to COLLECTIONS to include it)")
            continue
        ws = wb[name]
        rows = parse_tutorials(name, ws) if name == "TUTORIALS" else parse_default(name, ws)
        raw.extend(rows)
        print(f"  {name}: {len(rows)} links")

    sections, sec_lookup, items, seen = [], {}, [], set()
    for sheet, section, sub, name, url in raw:
        ci = col_idx[sheet]
        sec = clean_section(section)
        key = (ci, sec)
        if key not in sec_lookup:
            sec_lookup[key] = len(sections)
            sections.append({"c": ci, "n": sec})
        name, pick = clean_name(name)
        if not name:
            continue
        url = url.strip()
        if sub:
            sub = WS.sub(" ", FIRE.sub("", sub)).strip().rstrip(":").strip()
            if len(sub) > 60:
                sub = sub[:60].rsplit(" ", 1)[0] + "…"
        dedupe = (ci, name.lower(), url)
        if dedupe in seen:
            continue
        seen.add(dedupe)
        items.append([ci, sec_lookup[key], sub or "", name, url, domain(url), 1 if pick else 0])

    out = {
        "collections": [{"name": n, "code": c, "color": col} for _, n, c, col in COLLECTIONS],
        "sections": sections,
        "items": items,
    }
    data_js = "window.LIB=" + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";"
    if "</script" in data_js.lower():
        data_js = data_js.replace("</script", "<\\/script").replace("</SCRIPT", "<\\/SCRIPT")

    shell = open("shell.html", encoding="utf-8").read()
    open("index.html", "w", encoding="utf-8").write(shell.replace("__DATA__", data_js))
    print(f"Built index.html — {len(items)} resources, {len(sections)} sections, "
          f"{sum(i[6] for i in items)} staff picks")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else download()
    build(path)
